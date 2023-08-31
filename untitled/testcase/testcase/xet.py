
"""
import re
import requests
import json
import openpyxl


def write_excel(res):

    book = openpyxl.Workbook()  # 新建一个Excel
    sheet = book.create_sheet('导出数据',0)  # 创建sheet
    # title = ['姓名', '反馈信息', '手机号', '反馈时间']  # 写表头
    title = ['nickname', 'user_phone','content','adminmsg','id', 'app_id', 'avatar',  'replied_at', 'send_nick_name', 'user_id', 'type', 'wx_app_type', 'report_imgs', 'created_at']  # 写表头

    # 循环将表头写入到sheet页
    i = 1
    for header in title:
        sheet.cell(1, i, header)
        i += 1

    # # 写数据
    # for item in res:
        for row in range(0, len(res)):
            colRecord = res[row]
            for col in range(0, len(title)):
                # print(colRecord[title[col]])
                if title[col] == 'report_imgs' :
                    sheet.cell(row + 2, col + 1, ' '.join(colRecord[title[col]]))
                else :
                    sheet.cell(row + 2, col + 1, colRecord[title[col]])

    book.save("./用户反馈.xlsx")
    print("导出成功")



def xet_data():
    json_data = []
    page = 1
    while page <= 10:
        print(page)
        page += 1

        url = f"https://admin.xiaoe-tech.com/get/feedback_list? \
              ruler=content&search=&app_type=all&forbid=&page={page}"
        header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36',
            'Cookie': 'XIAOEID=8d1701a662d12cb8271119076777131a; mobile_manage=0; dataUpJssdkCookie={"wxver":"","net":"","sid":""}; Hm_lvt_081e3681cee6a2749a63db50a17625e2=1606186753; cookie_is_signed=1; channel=16-6820; cookie_channel=16-6820; b_user_token=token_5fbde2ef2f654LoIGb0T8ipJjoOJlyRdV; shop_type=lyRdV; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%22b_u_5ef327198485f_jEAVtzO0%22%2C%22first_id%22%3A%22175f830da523ef-091324ff7de698-74143944-1500000-175f830da53594%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%2C%22page_submodule%22%3A%22%E8%B4%A6%E5%8F%B7%E4%B8%BB%E9%A1%B5%22%2C%22page_name%22%3A%22%E8%B4%A6%E5%8F%B7%E4%B8%BB%E9%A1%B5_old%22%7D%2C%22%24device_id%22%3A%22175f830da523ef-091324ff7de698-74143944-1500000-175f830da53594%22%7D; appsc=appSHsLXnNy5032; with_app_id=appSHsLXnNy5032; cookie_session_id=cVusualJ2rYxWywLaK5frFEVCt7ZPINY; Hm_lpvt_081e3681cee6a2749a63db50a17625e2=1606280402; laravel_session=eyJpdiI6IkNJRE1qa3l6cUZFZEJYYUdLQXpac1E9PSIsInZhbHVlIjoiQjJycE42VWZESURXK3Z5UFI0cStTQUtNTXRqRGF5QWlJSHBXOVJPSHVqREE2VEUzNWFcL1UyZU9KckFSdXltTExBTk1LcnNIK2lqMEt3bnZhaG9EMG1nPT0iLCJtYWMiOiIxOWUwNGM1NDc3NTYzOGE0ZmU1MDhkNmE3Mjc0MmU3ZjRkNTk3NDQyMTNmNTg1YWFlM2Y2MTk0ZWY2NWU1Y2IxIn0%3D'}


        r = requests.get(url, headers=header)
        r.encoding = 'utf-8'
        # print (r.text)              # 获取网页的HTML字符串,该方式往往会出现乱码
        # print(r.status_code)       # 返回状态码
        # print(r.json())
        jsonobj = json.loads(r.text)
        print(jsonobj)
        json_data.extend(jsonobj['data'])
    print(json_data)
    write_excel(json_data)

if __name__ == '__main__':
    xet_data()


"""

