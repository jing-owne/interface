# time :2021/1/19 10:36
# Author :Gu xiao yu
import requests
import re
import json
#
# '''
# 基金代码、名称、简拼进行基金搜索
# '''
# search = '博时黄金'
# url = 'http://fundsuggest.eastmoney.com/FundSearch/api/FundSearchAPI.ashx?m=1&key=' + search  #
#
# result = requests.post(url)  # 发送请求
# print('##############查询结果##############')
# # print(result.text)  # 返回数据
#
# '''
# 通过基金编码获取估值
# '''
# code = '161028'
# url = 'http://fundgz.1234567.com.cn/js/%s.js' % code
# result1 = requests.get(url)  # 发送请求
# data = json.loads(re.match(".*?({.*}).*", result1.text, re.S).group(1))
# print('基金名称 {0}，估值增量 {1}'.format(data['name'],data['gszzl']))


import requests
import re
import json
from openpyxl import Workbook


class fund():
    def __init__(self, code):
        self.code = code

    def send(self):
        url = 'http://fundgz.1234567.com.cn/js/%s.js' % self.code
        result = requests.get(url)  # 发送请求
        data = json.loads(re.match(".*?({.*}).*", result.text, re.S).group(1))
        return '{0},{1}，涨幅 {2}'.format(i, data['name'], data['gszzl'])


# code是基金代码
code = [
    '005827',
    '161725',
    '160643',
    '161613',
    '004698',
    '001156',
    '240022',
    '161005',
    '163406',
    '260108',
    '003834',
    '001704',
    '001714',
    '001975',
    '004997',
    '006595',
    '001837',
    '110011',
    '110022',
    '003095'

]
for i in code:
    print(fund(i).send())

    #
    # def append_key(self):
    #     wb = Workbook()
    #     ws = wb.active
    #
    #     ws1 = wb.create_sheet("Mysheet")
    #     ws2 = wb.create_sheet("Mysheet", 0)
    #     ws3 = wb.create_sheet("Mysheet", -1)
    #     ws.title = "New Title"
    #     ws.sheet_properties.tabColor = "1072BA"
    #     ws3 = wb["New Title"]
    #     print(wb.sheetnames)
